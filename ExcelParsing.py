import openpyxl as xl;
from openpyxl import load_workbook
from openpyxl import Workbook
import re

class Main:

    def file_type(self):
        self._raw= self._wb1.cell(row=1,column=1)
        if self._raw.value==None:
            print("ISRACARD")
            self.isracard()
        else:
            print("VISA")
            self.visa()


    def read_file(self):
        self._fileName = "C:\OM\exampleVisaCard.xlsx"
        #self._fileName = "C:\OM\exampleMasterCard.xlsx"
        self._wb = load_workbook(self._fileName)
        self._wb1 = self._wb.active
        self._mr = self._wb1.max_row
        self._mc = self._wb1.max_column
        self.file_type()


    def visa(self):

        self._wb1.delete_cols(3)
        self._card = self._wb1.cell(row=2,column=1)
        self._s = str(self._card.value).split(",")[1].split(" ")[3]
        self._card_re = str(re.findall(r'[0-9]{4}', self._s))
        #self.output_file_visa()
        self.output_file()
        for i in range(4,self._mr):
            for j in range(1, 4):
                # reading cell value from source excel file
                self._c = self._wb1.cell(row=i, column=j)

                # writing the read value to destination excel file
                self._ws.cell(row=i - 2, column=j).value = self._c.value
                self._ws.cell(row=i - 2, column=5).value = self._card_re

        self._wb.save(filename="C:\OM\outputVisa2.xlsx")

    def isracard(self):
        self._wb1.delete_cols(3, 2)
        self._sheet=self._wb1
        self._alist=[]
        self._blist=[]
        self.output_file()
        for i in range(4, self._mr):
            self._b3 = self._sheet.cell(row=i, column=1)
            self._s = str(self._b3.value).split(" ")
            self._r = re.compile('[0-9]{4}')
            self._newlist = list(filter(self._r.match, self._s))
            self._list1 = [self._newlist, i]
            if self._newlist != []:
                self._alist.append(i)  # add row number for cars number to list
                self._blist.append(self._newlist)  # add cars number to list
            else:
                continue

        self._alist.append(self._mr + 5)

        for a in range(0, len(self._alist) - 1):
            for i in range(self._alist[a] + 3, self._alist[a + 1] - 2):
                for j in range(1, 4):
                    # reading cell value from source excel file
                    self._c = self._wb1.cell(row=i, column=j)
                    # writing the read value to destination excel file
                    self._ws.cell(row=i - 5, column=j).value = self._c.value
                    self._ws.cell(row=i - 5, column=5).value = str(self._blist[a])

        for j in range(1, 4):
            for i in range(1, self._mr):
                if self._ws.cell(row=i, column=1).value is None:
                    self._ws.delete_rows(i)
                else:
                    continue

        self._wb.save(filename="C:\OM\outputIsracard4.xlsx")


    def output_file(self):
        self._wb = Workbook()
        self._ws = self._wb.active
        self._x = ['תאריך עסקה', 'שם בית העסק', 'סכום עסקה', 'קטגוריה', 'מספר כרטיס']
        self._counter = 1
        for i in self._x:
            self._ws.cell(row=1, column=self._counter, value=i)
            self._counter += 1


Object = Main()
Object.read_file()