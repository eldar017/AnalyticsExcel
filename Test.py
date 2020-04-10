from openpyxl import Workbook
import json
import pika
from openpyxl import load_workbook
import pandas as pd
credentials = pika.PlainCredentials(username='user', password='password')
connection = pika.BlockingConnection(
    pika.ConnectionParameters(host='54.144.236.23', port='5672', credentials=credentials))
channel = connection.channel()

channel.queue_declare(queue='hello',durable=False)
counter = 1

def callback(ch, method, properties, body):
    print(" [x] Received %r" % body)
    data = str(body).encode('utf-8')
    #dd= pd.DataFrame(body)
    print(body)

    # data = str(body).split(",")
    # userID= data[2]
    # #print(userID)
    # excel = body
    # excel_dict = json.loads(excel)
    # #ecoded_response = response.read().decode("UTF-8")
    # print(json.dumps(excel_dict,indent=4,sort_keys=True))
    # excel_raw = excel_dict["jsonData"]
    # l = json.dumps(excel_raw,sort_keys=True,ensure_ascii=False).encode('utf-8')
    # print('===================================================================')
    # print(l)
    #
    # try:
    #     #data = pd.DataFrame(excel_raw)
    #     data = load_workbook(excel_raw)
    #     print(data)
    #     try:
    #         data.to_excel("c:\OM\pandas.xlsx")
    #     except:
    #         print('cannot write to file')
    # except:
    #     print('problem with read xls file')
    #
    #
    #data = pd.DataFrame(excel_raw)
    #print (excel_raw)
    # _wb = load_workbook(excel_raw)
    # _wb1 = _wb.active
    # print(_wb1)



channel.basic_consume(
    queue='hello', on_message_callback=callback, auto_ack=True)

print(' [*] Waiting for messages. To exit press CTRL+C')
channel.start_consuming()